from __future__ import annotations

import tempfile
import unittest
import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from class_schedule.class_model import HybridClass, NormalClass, Section
from class_schedule.config_schema import CatalogsFileSchema, CoursesFileSchema
from class_schedule.data_cleaning import (
    NORMALIZED_COLUMNS,
    clean_dataframe,
    clean_file,
    initialize_input,
)
from class_schedule.overrides import (
    OverrideEdit,
    OverrideFile,
    apply_overrides,
    load_overrides,
    locks_for_section,
    validate_override_context,
)
from class_schedule.schedule_model import ConstraintRule, Schedule, TimeWindow
from class_schedule.schedule_io import read_schedule
from class_schedule.solver import MeetingPattern, RoomRecord, SolverConfig, diff_schedules
from class_schedule.solver.candidates import section_candidates
from class_schedule.solver.config import (
    list_config_packages,
    resolve_config_paths,
    resolve_config_package,
)


class DataCleaningTests(unittest.TestCase):
    def test_normalizes_alias_columns_and_rejects_bad_rows(self):
        frame = pd.DataFrame([
            {
                "Subject": "math", "Number": "1113", "Section": "001",
                "Meeting_Days": "MWF", "Meeting_Times": "9:00 am-9:50 am",
                "Instructor Name": "Alice", "Room": "101", "Building": "Corley",
                "CRN": "12345", "Seats_Avail": "4",
            },
            {"Subject": "MATH", "Number": "", "Section": "002"},
        ])
        result = clean_dataframe(frame)
        self.assertEqual(tuple(result.normalized.columns), NORMALIZED_COLUMNS)
        self.assertEqual(result.normalized.iloc[0]["Time Slot"], "MWF 9:00am")
        self.assertEqual(result.normalized.iloc[0]["Delivery Mode"], "in_person")
        self.assertEqual(result.normalized.iloc[0]["CRN"], "12345")
        self.assertEqual(result.normalized.iloc[0]["Source Row"], 2)
        self.assertEqual(result.rejected.iloc[0]["Source Row"], 3)

    def test_clean_file_writes_auditable_bundle(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "raw.csv"
            pd.DataFrame([{
                "Subject": "MATH", "Number": "1113", "Section": "001",
                "Time Slot": "ONLINE", "Instructor": "Alice",
            }]).to_csv(source, index=False)
            destination = root / "normalized"
            clean_file(source, destination)
            self.assertTrue((destination / "sections.csv").is_file())
            self.assertTrue((destination / "rejected_rows.csv").is_file())
            self.assertTrue((destination / "validation.md").is_file())
            self.assertTrue((destination / "source_manifest.json").is_file())

    def test_initialize_writes_pre_change_views_beside_source(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "Course Schedule Report.csv"
            pd.DataFrame([
                {
                    "Subject": "MATH", "Number": "1113", "Section": "001",
                    "Time Slot": "MWF 9:00am", "Duration": "50",
                    "Instructor": "Original Instructor", "Room": "101",
                    "Building": "Corley",
                },
                {
                    "Subject": "STAT", "Number": "2163", "Section": "001",
                    "Time Slot": "ONLINE", "Instructor": "Online Instructor",
                },
            ]).to_csv(source, index=False)

            result = initialize_input(source, root / "normalized")

            self.assertEqual(result.draft_path, root / "draft" / "draft.csv")
            self.assertTrue(result.draft_path.is_file())

            instructor_path = root / "Course Schedule Report_instructor.xlsx"
            room_path = root / "Course Schedule Report_room.xlsx"
            self.assertEqual(result.instructor_path, instructor_path)
            self.assertEqual(result.room_path, room_path)
            workbook = load_workbook(instructor_path, read_only=True)
            try:
                self.assertEqual(
                    set(workbook.sheetnames),
                    {"Original Instructor", "Online Instructor"},
                )
            finally:
                workbook.close()
            workbook = load_workbook(room_path, read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Corley 101"])
            finally:
                workbook.close()

    def test_initialize_skips_views_when_a_source_row_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "raw.csv"
            pd.DataFrame([
                {"Subject": "MATH", "Number": "", "Section": "001"},
            ]).to_csv(source, index=False)

            result = initialize_input(source, root / "normalized")

            self.assertIsNone(result.instructor_path)
            self.assertIsNone(result.room_path)
            self.assertIsNone(result.draft_path)
            self.assertFalse((root / "raw_instructor.xlsx").exists())
            self.assertFalse((root / "raw_room.xlsx").exists())

    def test_known_cross_list_stays_unmarked_but_groups_during_validation(self):
        frame = pd.DataFrame([
            {"Subject": "MATH", "Number": "5173", "Section": "TC1",
             "Time Slot": "TBA", "Instructor": "Jordan, Scott M."},
            {"Subject": "STAT", "Number": "4173", "Section": "TC1",
             "Time Slot": "TBA", "Instructor": "Jordan, Scott M."},
        ])
        result = clean_dataframe(frame)
        self.assertEqual(len(set(result.normalized["Cross-List"])), 1)
        self.assertEqual(result.normalized.iloc[0]["Cross-List"], "")
        self.assertEqual(len(Schedule.from_dataframe(result.normalized)), 1)


class ScheduleIOTests(unittest.TestCase):
    def test_csv_becomes_a_grouped_schedule_at_the_file_boundary(self):
        frame = pd.DataFrame([
            {"Subject": "MATH", "Number": "5173", "Section": "TC1",
             "Time Slot": "TBA", "Instructor": "Jordan, Scott M."},
            {"Subject": "STAT", "Number": "4173", "Section": "TC1",
             "Time Slot": "TBA", "Instructor": "Jordan, Scott M."},
        ])
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "schedule.csv"
            frame.to_csv(path, index=False)
            schedule = read_schedule(path)

        self.assertEqual(len(schedule), 1)
        self.assertEqual(
            set(schedule.classes[0].course_ids),
            {"MATH 5173-TC1", "STAT 4173-TC1"},
        )


class OverrideTests(unittest.TestCase):
    def test_load_and_apply_edit_and_lock(self):
        schedule = Schedule([NormalClass((Section(
            "MATH", "1113", "001", "MWF 9:00am", 50,
            "101", "Alice", "Corley",
        ),))])
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "overrides.toml"
            path.write_text(
                '[[edits]]\ncourse_id = "MATH 1113-001"\n'
                'time_slot = "TR 9:30am"\nroom = "269"\nbuilding = "Corley"\n\n'
                '[[locks]]\ncourse_id = "MATH 1113-001"\n'
                'fields = ["time", "room", "building"]\n',
                encoding="utf-8",
            )
            overrides = load_overrides(path)
        changed = apply_overrides(schedule, overrides)
        section = changed.get("MATH 1113-001").sections[0]
        self.assertEqual(section.time_slot, "TR 9:30am")
        self.assertEqual(section.room, "269")
        self.assertEqual(
            locks_for_section(overrides.locks, ("MATH 1113-001",), 0),
            frozenset({"time", "room", "building"}),
        )

    def test_hybrid_edit_without_record_updates_physical_authority(self):
        hybrid = HybridClass((Section(
            "MATH", "1113", "F01", "MWF 9:00am", 50,
            "101", "Alice", "Corley",
        ),))
        schedule = Schedule([hybrid])
        overrides = OverrideFile(edits=(OverrideEdit(
            course_id="MATH 1113-F01",
            instructor="Bob",
            time_slot="TR 9:30am",
            room="269",
            building="Corley",
        ),))

        changed = apply_overrides(schedule, overrides)
        updated = changed.get("MATH 1113-F01")

        self.assertIsInstance(updated, HybridClass)
        self.assertEqual(updated.physical_section.time_slot, "TR 9:30am")
        self.assertEqual(updated.physical_section.room, "269")
        self.assertEqual(updated.physical_section.instructor, "Bob")
        self.assertEqual(updated.online_section.time_slot, "ONLINE")
        self.assertEqual(updated.online_section.room, "")
        self.assertEqual(updated.online_section.instructor, "Bob")

    def test_lock_filters_solver_candidates(self):
        section = Section(
            "MATH", "1113", "001", "MWF 9:00am", 50,
            "101", "Alice", "Corley",
        )
        config = SolverConfig(
            persons={}, preferences={},
            meeting_patterns=[MeetingPattern(
                "MWF", 50, (datetime.time(9), datetime.time(10)),
                frozenset({"normal"}),
            )],
            rooms=[RoomRecord("Corley", "101")],
        )
        candidates = section_candidates(
            NormalClass((section,)), section, config, 40, frozenset({"time"})
        )
        self.assertTrue(candidates)
        self.assertEqual({candidate.time_slot for candidate in candidates}, {"MWF 9:00am"})

    def test_configured_time_family_rejects_an_unlisted_current_start(self):
        section = Section(
            "MATH", "1003", "004", "MWF 11:50am", 50,
            "269", "Staff", "Corley",
        )
        config = SolverConfig(
            persons={}, preferences={},
            meeting_patterns=[MeetingPattern(
                "MWF", 50,
                (datetime.time(10), datetime.time(11), datetime.time(12)),
                frozenset({"normal"}),
            )],
            rooms=[RoomRecord("Corley", "269")],
            constraint_rules=(ConstraintRule(
                direction="-",
                time=TimeWindow(
                    frozenset("F"), datetime.time(12), datetime.time(12, 50)
                ),
            ),),
        )
        candidates = section_candidates(
            NormalClass((section,)), section, config, 40,
        )
        slots = {candidate.time_slot for candidate in candidates}
        self.assertNotIn("MWF 11:50am", slots)
        self.assertNotIn("MWF 12:00pm", slots)
        self.assertIn("MWF 10:00am", slots)

    def test_explicit_seminar_patterns_allow_single_weekdays_not_mwf(self):
        section = Section(
            "MATH", "4971", "001", "T 11:00am", 80,
            "101", "Alice", "Corley",
        )
        config = SolverConfig(
            persons={}, preferences={},
            meeting_patterns=[
                MeetingPattern(
                    days="MWF", duration_minutes=50,
                    starts=(datetime.time(9),),
                    roles=frozenset({"normal"}),
                ),
                MeetingPattern(
                    days="TR", duration_minutes=80,
                    starts=(datetime.time(11),),
                    roles=frozenset({"normal"}),
                ),
                MeetingPattern(
                    days="M", duration_minutes=50,
                    starts=(datetime.time(9),),
                    roles=frozenset({"normal"}),
                    courses=frozenset({"MATH 4971"}),
                ),
                MeetingPattern(
                    days="T", duration_minutes=80,
                    starts=(datetime.time(11),),
                    roles=frozenset({"normal"}),
                    courses=frozenset({"MATH 4971"}),
                ),
            ],
            rooms=[RoomRecord("Corley", "101")],
        )
        candidates = section_candidates(
            NormalClass((section,)), section, config, 40
        )
        slots = {candidate.time_slot for candidate in candidates}
        self.assertIn("T 11:00am", slots)
        self.assertIn("M 9:00am", slots)
        self.assertNotIn("MWF 9:00am", slots)
        self.assertNotIn("TR 11:00am", slots)
        durations = {
            (candidate.days, candidate.duration) for candidate in candidates
        }
        self.assertEqual(durations, {("M", 50), ("T", 80)})

    def test_version_bound_override_rejects_the_wrong_context(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "overrides.toml"
            path.write_text(
                'term = "27S"\nsource_version = "ver10"\n',
                encoding="utf-8",
            )
            overrides = load_overrides(path)
        validate_override_context(overrides, term="27S", source_version="ver10")
        with self.assertRaisesRegex(ValueError, "source_version"):
            validate_override_context(overrides, term="27S", source_version="ver9")
        with self.assertRaisesRegex(ValueError, "term"):
            validate_override_context(overrides, term="27F", source_version="ver10")


class ConfigLayoutTests(unittest.TestCase):
    @staticmethod
    def make_package(root: Path, name: str) -> Path:
        package = root / name
        (package / "basicinfo").mkdir(parents=True)
        for filename in ("catalogs.toml", "locations.toml", "timeslot.toml", "persons.toml"):
            (package / "basicinfo" / filename).write_text("", encoding="utf-8")
        for filename in ("courses.toml", "preferences.toml", "constraints.toml"):
            (package / filename).write_text("", encoding="utf-8")
        return package

    def test_resolves_the_fixed_seven_file_layout(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            package = self.make_package(root, "27S")
            paths = resolve_config_paths(root, "27S")
            self.assertEqual(len(paths), 7)
            self.assertEqual(paths["persons.toml"], package / "basicinfo" / "persons.toml")
            self.assertEqual(paths["courses.toml"], package / "courses.toml")

    def test_discovers_complete_packages_by_directory_name(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            package = self.make_package(root, "department_a")
            packages = list_config_packages(root)
            self.assertEqual([item.id for item in packages], ["department_a"])
            self.assertEqual(packages[0].display_name, "department_a")
            self.assertEqual(resolve_config_package(root, "department_a").root, package)

    def test_rejects_unknown_or_unsafe_package_names(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with self.assertRaisesRegex(FileNotFoundError, "Unknown"):
                resolve_config_package(root, "missing")
            with self.assertRaisesRegex(ValueError, "Invalid"):
                resolve_config_package(root, "../private")

    def test_courses_require_catalog_entries(self):
        catalogs = CatalogsFileSchema.model_validate({
            "courses": [{
                "subject": "MATH", "number": "1113",
                "title": "College Algebra", "credits": 3,
            }]
        })
        courses = CoursesFileSchema.model_validate({
            "courses": [{
                "subject": "MATH", "number": "1113", "sections": ["001"],
            }]
        })
        self.assertEqual(catalogs.courses[0].number, "1113")
        self.assertEqual(courses.courses[0].sections, ["001"])

    def test_relationship_rejects_an_unknown_section(self):
        with self.assertRaisesRegex(ValueError, "unknown sections"):
            CoursesFileSchema.model_validate({
                "courses": [{
                    "subject": "MATH", "number": "1113", "sections": ["001"],
                }],
                "relationships": [{
                    "id": "missing", "kind": "coreq",
                    "members": ["MATH 1113 001", "MATH 1110 001"],
                }],
            })

    def test_configured_coreq_uses_existing_atomic_class_behavior(self):
        relationships = CoursesFileSchema.model_validate({
            "courses": [
                {"subject": "MATH", "number": "2003", "sections": ["001"]},
                {"subject": "MATH", "number": "2004", "sections": ["001"]},
            ],
            "relationships": [{
                "id": "configured-coreq", "kind": "coreq",
                "members": ["MATH 2003 001", "MATH 2004 001"],
            }],
        }).relationships
        schedule = Schedule.from_records([
            {"Subject": "MATH", "Number": "2003", "Section": "001", "Time Slot": "MWF 9:00am", "Duration": 50, "Building": "Corley", "Room": "101", "Instructor": "Alice"},
            {"Subject": "MATH", "Number": "2004", "Section": "001", "Time Slot": "MWF 10:00am", "Duration": 50, "Building": "Corley", "Room": "101", "Instructor": "Alice"},
        ], relationships=relationships)
        self.assertEqual(type(schedule.classes[0]).__name__, "CoreqClass")


class CumulativeDiffTests(unittest.TestCase):
    def test_matches_by_course_identity_not_row_position(self):
        first = NormalClass((Section(
            "MATH", "1113", "001", "MWF 9:00am", 50,
            "101", "Alice", "Corley",
        ),))
        second = NormalClass((Section(
            "MATH", "2103", "002", "MWF 10:00am", 50,
            "102", "Bob", "Corley",
        ),))
        changed_first = NormalClass((Section(
            "MATH", "1113", "001", "MWF 11:00am", 50,
            "101", "Alice", "Corley",
        ),))
        changes = diff_schedules(
            Schedule([first, second]), Schedule([second, changed_first])
        )
        self.assertEqual(
            [(change.course_id, change.field) for change in changes],
            [("MATH 1113-001", "time")],
        )

    def test_reports_added_and_removed_sections(self):
        before = Schedule([NormalClass((Section(
            "MATH", "1113", "001", "ONLINE", None, "", "Alice",
        ),))])
        after = Schedule([NormalClass((Section(
            "MATH", "2103", "002", "ONLINE", None, "", "Bob",
        ),))])
        changes = diff_schedules(before, after)
        self.assertEqual(
            [(change.course_id, change.field, change.after) for change in changes],
            [
                ("MATH 1113-001", "status", "removed"),
                ("MATH 2103-002", "status", "added"),
            ],
        )


if __name__ == "__main__":
    unittest.main()
