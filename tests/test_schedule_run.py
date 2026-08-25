import tempfile
import unittest
import json
import hashlib
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from class_schedule.schedule_run import (
    create_override_template,
    infer_parent_version,
    install_version_override_template,
    latest_version,
    next_version,
    publish_final,
    run_term,
    version_schedule_path,
    worst_overload,
)
from class_schedule.schedule_io import read_schedule
from class_schedule.schedule_model import PersonRecord, Schedule
from class_schedule.solver import SolverConfig


class VersionTests(unittest.TestCase):
    def test_worst_overload_excludes_the_configured_tolerance(self):
        schedule = Schedule.from_records([{
            "Subject": "MATH", "Number": "1113", "Section": "001",
            "Instructor": "Alice", "Time Slot": "MWF 9:00am",
            "Duration": 50, "Room": "101", "Building": "Corley",
            "Credits": 13,
        }])
        config = SolverConfig(
            persons={"Alice": PersonRecord(name="Alice", max_load=12)},
            preferences={}, meeting_patterns=[], rooms=[],
        )

        self.assertEqual(worst_overload(schedule, config), 0)

    def test_next_version_ignores_unrelated_entries(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            (root / "ver1").mkdir()
            (root / "ver4").mkdir()
            (root / "version9").mkdir()
            (root / "ver4_validation").mkdir()
            (root / "final").mkdir()
            self.assertEqual(latest_version(root), "ver4")
            self.assertEqual(next_version(root), "ver5")

    def test_infers_parent_only_from_an_exact_version_directory(self):
        root = Path("out") / "27S"
        self.assertEqual(
            infer_parent_version(root / "ver7" / "27S_ver7.csv", root), "ver7"
        )
        self.assertIsNone(
            infer_parent_version(root / "ver7_archive" / "27S_ver7.csv", root)
        )

    def test_resolves_only_a_canonical_published_schedule(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            path = root / "27S" / "ver3" / "27S_ver3.csv"
            path.parent.mkdir(parents=True)
            path.write_text("Subject,Number,Section\n", encoding="utf-8")
            self.assertEqual(
                version_schedule_path("27S", "ver3", output_root=root), path
            )
            with self.assertRaises(ValueError):
                version_schedule_path("27S", "3", output_root=root)


class RunTermTests(unittest.TestCase):
    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    @staticmethod
    def _write_source(
        path: Path,
        *,
        time_slot: str = "MWF 9:00am",
        room: str = "101",
    ) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame([{
            "Subject": "MATH",
            "Number": "1113",
            "Section": "001",
            "Instructor": "Bain, Leslie M.",
            "Time Slot": time_slot,
            "Duration": 50,
            "Room": room,
            "Building": "Corley",
        }]).to_csv(path, index=False)

    def _write_initial_artifact(self, path: Path) -> Path:
        self._write_source(path)
        changes = path.parent / "changes.toml"
        changes.write_text("# no term changes\n", encoding="utf-8")
        (path.parent / "manifest.json").write_text(
            json.dumps({
                "schema_version": 1,
                "role": "initial",
                "changes": {"path": str(changes), "sha256": self._sha256(changes)},
                "initial": {"path": path.name, "sha256": self._sha256(path)},
            }),
            encoding="utf-8",
        )
        return changes

    def _write_version_provenance(self, version_dir: Path) -> None:
        baseline = version_dir / "baseline.csv"
        changes = version_dir / "applied_changes.toml"
        changes.write_text("# initial changes snapshot\n", encoding="utf-8")
        (version_dir / "manifest.json").write_text(
            json.dumps({
                "initial_baseline": {
                    "role": "initial", "snapshot": baseline.name,
                    "path": str(baseline), "sha256": self._sha256(baseline),
                },
                "files": {baseline.name: self._sha256(baseline)},
            }),
            encoding="utf-8",
        )

    def test_writes_atomic_csv_report_and_attempt_summary(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            input_path = root / "work" / "TEST" / "initial" / "initial.csv"
            self._write_initial_artifact(input_path)

            bundle = run_term(
                "TEST",
                input_path=input_path,
                initial_path=input_path,
                output_root=root / "out",
                config_dir="config",
                attempts=3,
                time_limit_seconds=5,
            )

            self.assertEqual(bundle.version, "ver1")
            self.assertTrue(bundle.schedule_path.exists())
            self.assertTrue(bundle.instructor_path.exists())
            self.assertTrue(bundle.room_path.exists())
            self.assertTrue(bundle.report_path.exists())
            self.assertTrue(bundle.attempts_path.exists())
            self.assertTrue(bundle.changes_path.exists())
            self.assertTrue(bundle.baseline_path.exists())
            self.assertEqual(bundle.baseline_path.read_bytes(), input_path.read_bytes())
            self.assertTrue(bundle.manifest_path.exists())
            self.assertTrue(bundle.overrides_path.exists())
            self.assertTrue(bundle.applied_overrides_path.exists())
            self.assertTrue(bundle.applied_changes_path.exists())
            manifest = json.loads(bundle.manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["schema_version"], 4)
            self.assertEqual(manifest["solver"]["search_workers"], 8)
            self.assertEqual(manifest["solver"]["attempts_requested"], 3)
            self.assertEqual(manifest["solver"]["attempts_run"], 1)
            self.assertEqual(len(bundle.attempts), 1)
            self.assertEqual(manifest["initial_baseline"]["path"], str(input_path))
            self.assertEqual(manifest["initial_baseline"]["snapshot"], "baseline.csv")
            self.assertIn("baseline.csv", manifest["files"])
            self.assertIn("TEST_ver1_instructor.xlsx", manifest["files"])
            self.assertIn("TEST_ver1_room.xlsx", manifest["files"])
            self.assertIn("applied_overrides.toml", manifest["files"])
            self.assertIn("applied_changes.toml", manifest["files"])
            self.assertNotIn("overrides.toml", manifest["files"])
            self.assertTrue(manifest["override_workspace"]["mutable"])
            self.assertIn("source_version = \"ver1\"", bundle.overrides_path.read_text())
            written = pd.read_csv(bundle.schedule_path, dtype=str)
            instructor = written.loc[0, "Instructor"]
            room = f'{written.loc[0, "Building"]} {written.loc[0, "Room"]}'
            workbook = load_workbook(bundle.instructor_path, read_only=True)
            try:
                self.assertIn(instructor, workbook.sheetnames)
            finally:
                workbook.close()
            workbook = load_workbook(bundle.room_path, read_only=True)
            try:
                self.assertIn(room, workbook.sheetnames)
            finally:
                workbook.close()
            report = bundle.report_path.read_text(encoding="utf-8")
            self.assertIn("Configuration version", report)
            self.assertIn("## Attempt comparison", report)
            self.assertIn("## Remaining soft findings", report)
            self.assertNotEqual(written.loc[0, "Instructor"], "Bain")

            second = run_term(
                "TEST", input_path=input_path, initial_path=input_path,
                output_root=root / "out", config_dir="config",
                attempts=1, time_limit_seconds=5,
            )
            second_manifest = json.loads(
                second.manifest_path.read_text(encoding="utf-8")
            )
            self.assertEqual(second.version, "ver2")
            self.assertIsNone(second_manifest["parent"])
            self.assertEqual(second_manifest["input"]["path"], str(input_path))
            self.assertEqual(second_manifest["initial_baseline"]["role"], "initial")
            self.assertEqual(
                second.baseline_path.read_bytes(), input_path.read_bytes()
            )

            with self.assertRaisesRegex(ValueError, "must start from initial"):
                run_term(
                    "TEST", input_path=bundle.schedule_path, parent="ver1",
                    output_root=root / "out", config_dir="config",
                    attempts=1, time_limit_seconds=5,
                )

    def test_solve_rejects_cancelled_course_instead_of_mutating_initial(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            input_path = root / "previous_version.csv"
            pd.DataFrame([
                {
                    "Subject": "MATH", "Number": "1113", "Section": "001",
                    "Instructor": "Bain, Leslie M.", "Time Slot": "MWF 9:00am",
                    "Duration": 50, "Room": "101", "Building": "Corley",
                },
                {
                    "Subject": "MATH", "Number": "4993", "Section": "001",
                    "Instructor": "Limperis, Thomas G.", "Time Slot": "",
                    "Duration": "", "Room": "", "Building": "",
                },
            ]).to_csv(input_path, index=False)
            changes_path = root / "changes.toml"
            changes_path.write_text(
                '[[cancel_courses]]\nsubject = "MATH"\nnumber = "4993"\n',
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "rebuild initial"):
                run_term(
                    "TEST", input_path=input_path, baseline_path=input_path,
                    changes_path=changes_path, output_root=root / "out",
                    config_dir="config", attempts=1, time_limit_seconds=5,
                    historical_backfill=True,
                )

    def test_refuses_to_overwrite_an_existing_version(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            destination = root / "out" / "TEST" / "ver2"
            destination.mkdir(parents=True)
            with self.assertRaises(FileExistsError):
                run_term(
                    "TEST",
                    input_path=root / "missing.csv",
                    output_root=root / "out",
                    version="ver2",
                    attempts=1,
                )

    def test_generates_a_parseable_template_bound_to_the_source_version(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "out" / "TEST" / "ver3" / "TEST_ver3.csv"
            self._write_source(source)
            destination = root / "revision.toml"
            written = create_override_template(
                "TEST", "ver3", output_path=destination,
                output_root=root / "out", config_dir="config",
            )
            text = written.read_text(encoding="utf-8")
            self.assertIn('term = "TEST"', text)
            self.assertIn('source_version = "ver3"', text)
            self.assertIn("MATH 1113-001 | record=0", text)

    def test_embedded_override_refreshes_final_without_creating_a_version(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            output_root = root / "out"
            source = output_root / "TEST" / "ver3" / "TEST_ver3.csv"
            self._write_source(source, time_slot="MWF 10:00am")
            self._write_source(source.parent / "baseline.csv")
            self._write_version_provenance(source.parent)
            overrides = source.parent / "overrides.toml"
            overrides.write_text(
                'term = "TEST"\nsource_version = "ver3"\n\n'
                '[[edits]]\ncourse_id = "MATH 1113-001"\n'
                'instructor = "Bain, Leslie M."\n'
                'time_slot = "MWF 9:00am"\nroom = "102"\nbuilding = "Corley"\n\n'
                '[[locks]]\ncourse_id = "MATH 1113-001"\n'
                'fields = ["instructor", "time", "room", "building"]\n',
                encoding="utf-8",
            )

            bundle = publish_final(
                "TEST", "ver3",
                output_root=output_root, config_dir="config",
                attempts=1, time_limit_seconds=5,
            )

            self.assertEqual(bundle.version, "final")
            revised = read_schedule(bundle.schedule_path)
            self.assertEqual(
                revised.get("MATH 1113-001").sections[0].time_slot,
                "MWF 9:00am",
            )
            self.assertEqual(
                revised.get("MATH 1113-001").sections[0].room, "102"
            )
            cumulative = pd.read_csv(bundle.changes_path, dtype=str)
            self.assertNotIn("time", set(cumulative["Field"]))
            self.assertEqual(list(cumulative["Field"]), ["room"])
            manifest = json.loads(bundle.manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["parent"], "ver3")
            self.assertEqual(manifest["input"]["path"], str(source))
            self.assertFalse(manifest["override_workspace"]["mutable"])
            self.assertIn("overrides.toml", manifest["files"])
            self.assertIn("applied_overrides.toml", manifest["files"])
            self.assertEqual(bundle.schedule_path.name, "TEST_final.csv")

            overrides.write_text(
                overrides.read_text(encoding="utf-8").replace(
                    'room = "102"', 'room = "103"'
                ),
                encoding="utf-8",
            )
            refreshed = publish_final(
                "TEST", "ver3",
                output_root=output_root, config_dir="config",
                attempts=1, time_limit_seconds=5,
            )
            schedule = read_schedule(refreshed.schedule_path)
            self.assertEqual(
                schedule.get("MATH 1113-001").sections[0].room,
                "103",
            )
            self.assertFalse((output_root / "TEST" / "ver4").exists())

    def test_final_rejects_an_untouched_override_template(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            output_root = root / "out"
            source = output_root / "TEST" / "ver3" / "TEST_ver3.csv"
            self._write_source(source)
            (source.parent / "overrides.toml").write_text(
                'term = "TEST"\nsource_version = "ver3"\n',
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "No manual edits or locks"):
                publish_final(
                    "TEST", "ver3", output_root=output_root,
                    config_dir="config", attempts=1, time_limit_seconds=5,
                )
            self.assertFalse((output_root / "TEST" / "final").exists())

    def test_installs_an_embedded_workspace_and_preserves_the_old_override(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "out" / "TEST" / "ver3" / "TEST_ver3.csv"
            self._write_source(source)
            old = source.parent / "overrides.toml"
            old.write_text("# old applied input\n", encoding="utf-8")
            workspace = install_version_override_template(
                "TEST", "ver3", output_root=root / "out", config_dir="config"
            )
            self.assertIn('source_version = "ver3"', workspace.read_text())
            self.assertEqual(
                (source.parent / "applied_overrides.toml").read_text(),
                "# old applied input\n",
            )


if __name__ == "__main__":
    unittest.main()
