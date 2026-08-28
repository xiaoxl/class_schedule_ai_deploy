"""Schedule: an in-memory collection of atomic classes.

This module owns everything ``class_model`` deliberately does not: grouping
many table records into atomic classes, editing and evaluating the grouped
collection, and serializing it to/from a pandas DataFrame. Disk access remains
in ``schedule_io``. This module composes ``class_model``'s public API --
``Section``, the ``Class`` hierarchy, and each kind's recognition predicate
(``is_four_credit``, ``is_hybrid``, ``is_cross_listing``, ``is_coreq_pair``)
-- without re-implementing any of their rules. Editing a class's time, room,
or instructor is delegated straight to that class's own ``change_*`` method,
so kind-specific behavior (``FourCreditClass``'s day-pattern-aware
``change_time``, ``CoreqClass`` requiring an explicit record, ...) applies
automatically -- this module doesn't need to know about any of it.
"""

from __future__ import annotations

import datetime
import re
import tomllib
from collections.abc import Iterable, Iterator, Mapping
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import record_utils
from .config_schema import (
    CatalogCourseSchema,
    BackToBackPolicySchema,
    CourseRelationshipSchema,
    FlatPreferenceRuleSchema,
    NewInstructorPolicySchema,
    NewProfessorPolicySchema,
    PersonsFileSchema,
    PreferencesFileSchema,
    TimeWindowSchema,
    WorkloadPolicySchema,
)
from .class_model import (
    Class,
    CoreqClass,
    CrossListingClass,
    FourCreditClass,
    HybridClass,
    NormalClass,
    Section,
)
from .instructor_identity import is_new_instructor, is_new_professor
from .pattern_rules import MeetingPatternLike, matches_configured_pattern

_UNSET = object()

# Concurrent (dual-credit/high-school partnership) sections -- these are
# never real classes to schedule, so they're dropped before any grouping
# even sees them, not just excluded from one specific pairing rule.
_IGNORED_SECTION_PREFIXES = ("P", "ET", "A")


class GroupingError(ValueError):
    """Raised when a group of CSV rows can't become a valid class.

    Carries the raw row dicts involved (``records``), not just a message,
    so a caller -- e.g. the web app -- can show the user exactly which
    rows are implicated instead of only a course id buried in text.
    """

    def __init__(self, message: str, records: list[dict]) -> None:
        super().__init__(message)
        self.records = records


class Schedule:
    """An editable collection of atomic classes."""

    def __init__(self, classes: Iterable[Class] = ()) -> None:
        self.classes: list[Class] = list(classes)

    # ---- import (CSV records/DataFrame -> Schedule) ----

    @classmethod
    def from_records(
        cls,
        records: Iterable[Mapping[str, object]],
        *,
        persons: Mapping[str, "PersonRecord"] | None = None,
        relationships: Iterable[CourseRelationshipSchema] = (),
        catalogs: Iterable[CatalogCourseSchema] = (),
        infer_legacy_relationships: bool = True,
        infer_marked_cross_lists: bool = False,
    ) -> "Schedule":
        """Group a complete table of CSV records into atomic classes."""
        return cls(_group_records(
            records, persons=persons, relationships=relationships, catalogs=catalogs,
            infer_legacy_relationships=infer_legacy_relationships,
            infer_marked_cross_lists=infer_marked_cross_lists,
        ))

    @classmethod
    def from_dataframe(
        cls,
        dataframe: pd.DataFrame,
        *,
        persons: Mapping[str, "PersonRecord"] | None = None,
        relationships: Iterable[CourseRelationshipSchema] = (),
        catalogs: Iterable[CatalogCourseSchema] = (),
        infer_legacy_relationships: bool = True,
        infer_marked_cross_lists: bool = False,
    ) -> "Schedule":
        """Group a complete DataFrame into atomic classes."""
        return cls.from_records(
            dataframe.to_dict(orient="records"), persons=persons,
            relationships=relationships,
            catalogs=catalogs,
            infer_legacy_relationships=infer_legacy_relationships,
            infer_marked_cross_lists=infer_marked_cross_lists,
        )

    # ---- export (Schedule -> CSV records/DataFrame) ----

    def to_records(self) -> list[dict[str, object]]:
        """Flatten every class back into one list of CSV row dicts."""
        return [
            record for item in self.classes for record in item.to_records()
        ]

    def to_dataframe(self) -> pd.DataFrame:
        """Flatten every class into one DataFrame."""
        records = self.to_records()
        if not records:
            return pd.DataFrame()
        return pd.DataFrame.from_records(records)

    # ---- export (Schedule -> Excel workbooks) ----
    #
    # Three operational views, each its own file: the flat normalized
    # table, and two weekly-grid workbooks (one worksheet per instructor,
    # one per room). All three still export even when the schedule has
    # conflicts -- see ``_build_weekly_sheet``.

    def to_raw_excel(self, path: str | Path) -> None:
        """One row per CSV record, in the schedule's own column order."""
        _write_raw_excel(self.to_dataframe(), path)

    def to_instructor_excel(self, path: str | Path) -> None:
        """One worksheet per instructor: a Monday-Friday weekly grid."""
        _weekly_workbook(self, group="instructor").save(path)

    def to_room_excel(self, path: str | Path) -> None:
        """One worksheet per room: a Monday-Friday weekly grid."""
        _weekly_workbook(self, group="room").save(path)

    # ---- lookup ----

    def __iter__(self):
        return iter(self.classes)

    def __len__(self) -> int:
        return len(self.classes)

    @property
    def course_ids(self) -> list[str]:
        return [
            course_id
            for item in self.classes
            for course_id in item.course_ids
        ]

    def index_of(self, course_id: str) -> int:
        """Return the index of the class containing ``course_id``."""
        for index, item in enumerate(self.classes):
            if course_id in item.course_ids:
                return index
        raise KeyError(f"Course not found: {course_id}")

    def get(self, course_id: str) -> Class:
        return self.classes[self.index_of(course_id)]

    # ---- collection editing ----

    def add(self, item: Class) -> "Schedule":
        """Add a class, rejecting any course id already present."""
        clashing = set(self.course_ids).intersection(item.course_ids)
        if clashing:
            raise ValueError(f"Course already exists: {sorted(clashing)}")
        self.classes.append(item)
        return self

    def remove(self, course_id: str) -> "Schedule":
        del self.classes[self.index_of(course_id)]
        return self

    # ---- per-class editing: delegates to that class's own change_* ----
    #
    # ``record`` defaults to ``_UNSET`` rather than ``None`` so that an
    # omitted ``record`` is forwarded as *omitted*, not as an explicit
    # ``None``. Some kinds (``CoreqClass``) make ``record`` a required
    # keyword with no default of its own, specifically so a caller must
    # choose which meeting to move; if this wrapper always forwarded
    # ``record=None`` it would silently satisfy that requirement and
    # defeat the whole point.

    def change_time(
        self, course_id: str, time_slot: str, *, record: object = _UNSET
    ) -> "Schedule":
        index = self.index_of(course_id)
        kwargs = {} if record is _UNSET else {"record": record}
        self.classes[index] = self.classes[index].change_time(
            time_slot, **kwargs
        )
        return self

    def change_room(
        self, course_id: str, room: str, *, record: object = _UNSET
    ) -> "Schedule":
        index = self.index_of(course_id)
        kwargs = {} if record is _UNSET else {"record": record}
        self.classes[index] = self.classes[index].change_room(
            room, **kwargs
        )
        return self

    def change_instructor(
        self, course_id: str, instructor: str, *, record: object = _UNSET
    ) -> "Schedule":
        index = self.index_of(course_id)
        kwargs = {} if record is _UNSET else {"record": record}
        self.classes[index] = self.classes[index].change_instructor(
            instructor, **kwargs
        )
        return self


def _group_records(
    records: Iterable[Mapping[str, object]],
    *,
    persons: Mapping[str, "PersonRecord"] | None = None,
    relationships: Iterable[CourseRelationshipSchema] = (),
    catalogs: Iterable[CatalogCourseSchema] = (),
    infer_legacy_relationships: bool = True,
    infer_marked_cross_lists: bool = False,
) -> list[Class]:
    """Group raw CSV records into atomic classes.

    Classification happens as part of the scan itself -- there is no
    separate "infer" step. Each pass below both finds candidate rows
    *and* decides which kind they become, using that kind's own public
    recognition predicate from ``class_model``. Precedence, highest
    first: same-course pairs (four-credit/Hybrid) beat cross-listed
    pairs, which beat coreq pairs. A remaining physical M/F-prefixed row is
    normalized into a Hybrid with a derived ONLINE companion; anything else
    left over is a single class.
    Construction always re-validates through the target class's own
    ``validate`` -- a pair that matches a scan (e.g. the coreq whitelist)
    but fails its finer rules (e.g. bad scheduling) still raises here.

    Rows in a concurrent-enrollment section (see
    ``_IGNORED_SECTION_PREFIXES``) are dropped up front and never appear
    in the resulting classes.
    """
    catalog_by_id = {(item.subject, item.number): item for item in catalogs}
    remaining: list[Section] = []
    for row in records:
        normalized = record_utils.normalize_columns(row)
        section_code = record_utils.text(
            record_utils.value(normalized, "Section")
        ).upper()
        if section_code.startswith(_IGNORED_SECTION_PREFIXES):
            continue
        subject = record_utils.text(record_utils.value(normalized, "Subject")).upper()
        number = record_utils.text(record_utils.value(normalized, "Number")).upper()
        catalog = catalog_by_id.get((subject, number))
        if catalog_by_id and catalog is None:
            raise GroupingError(
                f"{subject} {number} is missing from catalogs.toml",
                [dict(normalized)],
            )
        if catalog is not None:
            raw_credits = record_utils.text(record_utils.value(normalized, "Credits"))
            if raw_credits:
                try:
                    input_credits = float(raw_credits)
                except ValueError as error:
                    raise GroupingError(
                        f"{subject} {number} has invalid input Credits={raw_credits!r}",
                        [dict(normalized)],
                    ) from error
                if abs(input_credits - catalog.credits) > 1e-9:
                    raise GroupingError(
                        f"{subject} {number} has input Credits={input_credits:g}, "
                        f"but catalogs.toml declares {catalog.credits:g}",
                        [dict(normalized)],
                    )
            normalized["Title"] = catalog.title
            normalized["Credits"] = catalog.credits
        if record_utils.text(normalized.get("Cross-List")).startswith("configured:"):
            # Normalize the supported ``configured:`` compatibility marker;
            # known pairs are recognized directly by CrossListingClass.
            normalized["Cross-List"] = ""
        if persons:
            instructor = record_utils.text(record_utils.value(normalized, "Instructor"))
            resolved = resolve_person_name(instructor, persons, subject=subject)
            if resolved is not None:
                normalized["Instructor"] = resolved
        try:
            remaining.append(Section.from_record(normalized))
        except ValueError as error:
            raise GroupingError(str(error), [dict(normalized)]) from error

    result, remaining = _take_configured_relationships(
        remaining, tuple(relationships)
    )
    same_course, remaining = _take_same_course(remaining)
    result.extend(same_course)
    if infer_legacy_relationships:
        cross_listed, remaining = _take_cross_listed(remaining)
        result.extend(cross_listed)
        coreqs, remaining = _take_coreqs(remaining)
        result.extend(coreqs)
    elif infer_marked_cross_lists:
        cross_listed, remaining = _take_cross_list_column(remaining)
        result.extend(cross_listed)
    result.extend(
        HybridClass((section,))
        if HybridClass.is_hybrid_physical(section)
        else NormalClass((section,))
        for section in remaining
    )
    return result


def _take_configured_relationships(
    remaining: list[Section],
    relationships: tuple[CourseRelationshipSchema, ...],
) -> tuple[list[Class], list[Section]]:
    """Apply explicit courses.toml relationships before legacy inference."""
    found: list[Class] = []
    consumed: set[int] = set()
    for relationship in relationships:
        groups = []
        for member in relationship.members:
            subject, number, section = member.split(maxsplit=2)
            matches = [
                item for item in remaining
                if item.subject == subject and item.number == number
                and item.section.upper() == section
                and id(item) not in consumed
            ]
            if not matches:
                continue  # courses.toml may include sections absent from an imported template
            groups.append(matches)
        if len(groups) != len(relationship.members):
            continue
        rows = tuple(item for group in groups for item in group)
        try:
            if relationship.kind == "hybrid":
                item = HybridClass(rows)
            elif relationship.kind == "four_credit":
                item = FourCreditClass(rows)
            elif relationship.kind == "cross_listing":
                if len(rows) != 2:
                    raise ValueError("configured cross-listing requires two source rows")
                synced_fields = (
                    frozenset(relationship.synced_fields)
                    if relationship.synced_fields is not None else None
                )
                item = CrossListingClass.from_configured_sections(
                    rows, synced_fields=synced_fields,
                )
            else:
                if len(rows) != 2:
                    raise ValueError("configured coreq requires two source rows")
                item = CoreqClass.from_configured_sections(rows)
        except ValueError as error:
            raise GroupingError(str(error), [row.to_record() for row in rows]) from error
        found.append(item)
        consumed.update(id(row) for row in rows)
    return found, [item for item in remaining if id(item) not in consumed]


def _take_same_course(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    by_identity: dict[tuple[str, str, str], list[Section]] = {}
    for section in remaining:
        by_identity.setdefault(section.identity, []).append(section)

    found: list[Class] = []
    consumed: set[int] = set()
    for key, group in by_identity.items():
        if len(group) > 2:
            raise GroupingError(
                f"{' '.join(key[:2])}-{key[2]} has more than two CSV records",
                [section.to_record() for section in group],
            )
        if len(group) == 2:
            left, right = group
            target = (
                HybridClass
                if HybridClass.is_hybrid(left, right)
                else FourCreditClass
            )
            try:
                found.append(target((left, right)))
            except ValueError as error:
                raise GroupingError(
                    str(error), [left.to_record(), right.to_record()]
                ) from error
            consumed.update(id(section) for section in group)
    return found, [
        section for section in remaining if id(section) not in consumed
    ]


def _take_cross_listed(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    found, remaining = _take_cross_list_column(remaining)
    more, remaining = _take_known_cross_list_pairs(remaining)
    found.extend(more)
    more, remaining = _take_honors_pairs(remaining)
    found.extend(more)
    return found, remaining


def _take_cross_list_column(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    by_cross_list: dict[str, list[Section]] = {}
    for section in remaining:
        if section.cross_list:
            by_cross_list.setdefault(section.cross_list, []).append(section)

    found: list[Class] = []
    consumed: set[int] = set()
    for cross_list, group in by_cross_list.items():
        if len(group) > 2:
            raise GroupingError(
                f"Cross-List {cross_list!r} has more than two CSV records",
                [section.to_record() for section in group],
            )
        if len(group) == 2:
            left, right = group
            try:
                found.append(CrossListingClass(tuple(group)))
            except ValueError as error:
                raise GroupingError(
                    str(error), [left.to_record(), right.to_record()]
                ) from error
            consumed.update(id(section) for section in group)
    return found, [
        section for section in remaining if id(section) not in consumed
    ]


def _take_known_cross_list_pairs(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    """Group built-in cross-list course pairs without a source marker."""
    matches: list[tuple[int, int]] = []
    for left_index, left in enumerate(remaining):
        for right_index in range(left_index + 1, len(remaining)):
            if CrossListingClass.is_known_pair(left, remaining[right_index]):
                matches.append((left_index, right_index))

    involved = [index for pair in matches for index in pair]
    if len(involved) != len(set(involved)):
        raise GroupingError(
            "Ambiguous known cross-list pairing: one course matches multiple pairs",
            [remaining[index].to_record() for index in sorted(set(involved))],
        )

    consumed = set(involved)
    found = [
        CrossListingClass((remaining[left], remaining[right]))
        for left, right in matches
    ]
    return found, [
        section for index, section in enumerate(remaining)
        if index not in consumed
    ]


def _take_honors_pairs(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    """Pair regular/'H'-prefixed honors sections not already linked by an
    explicit Cross-List value (see ``CrossListingClass.is_honors_pair``)."""
    matches: list[tuple[int, int]] = []
    for left_index, left in enumerate(remaining):
        for right_index in range(left_index + 1, len(remaining)):
            right = remaining[right_index]
            if CrossListingClass.is_honors_pair(left, right):
                matches.append((left_index, right_index))

    involved = [index for pair in matches for index in pair]
    if len(involved) != len(set(involved)):
        raise GroupingError(
            "Ambiguous honors-section pairing: one course matches "
            "multiple pairs",
            [remaining[index].to_record() for index in sorted(set(involved))],
        )

    consumed = set(involved)
    found = []
    for left_index, right_index in matches:
        left, right = remaining[left_index], remaining[right_index]
        try:
            found.append(CrossListingClass((left, right)))
        except ValueError as error:
            raise GroupingError(
                str(error), [left.to_record(), right.to_record()]
            ) from error
    return found, [
        section
        for index, section in enumerate(remaining)
        if index not in consumed
    ]


def _take_coreqs(
    remaining: list[Section],
) -> tuple[list[Class], list[Section]]:
    matches: list[tuple[int, int]] = []
    for left_index, left in enumerate(remaining):
        for right_index in range(left_index + 1, len(remaining)):
            right = remaining[right_index]
            if CoreqClass.is_coreq_pair(left, right):
                matches.append((left_index, right_index))

    involved = [index for pair in matches for index in pair]
    if len(involved) != len(set(involved)):
        raise GroupingError(
            "Ambiguous coreq records: one course matches multiple pairs",
            [remaining[index].to_record() for index in sorted(set(involved))],
        )

    consumed = set(involved)
    found = []
    for left_index, right_index in matches:
        left, right = remaining[left_index], remaining[right_index]
        try:
            found.append(CoreqClass((left, right)))
        except ValueError as error:
            raise GroupingError(
                str(error), [left.to_record(), right.to_record()]
            ) from error
    return found, [
        section
        for index, section in enumerate(remaining)
        if index not in consumed
    ]


# ---- configuration-driven evaluation (persons.toml and preferences.toml) ----
#
# persons.toml holds contractual identity, qualification, alias, and load
# facts; preferences.toml holds this term's wishes for an instructor.
# ``evaluate_schedule`` gets hard violations from structural double-booking
# and configured hard rules. Atomic-class legality is enforced at
# ``Class`` construction and by the solver's pairwise validity constraints,
# so invalid four-credit/Hybrid/cross-list/coreq combinations cannot become
# a Schedule or a solved result.
#
# max_load is a target, not just a ceiling. All workload thresholds and
# penalties are supplied by constraints.toml through WorkloadPolicySchema.


def weekday_time_overlap(
    days_a: str | Iterable[str] | None,
    start_a: datetime.time | None,
    end_a: datetime.time | None,
    days_b: str | Iterable[str] | None,
    start_b: datetime.time | None,
    end_b: datetime.time | None,
) -> bool:
    """Shared by ``TimeWindow.overlaps`` and ``overlaps_in_time`` -- true
    when both sides meet on at least one common weekday AND their clock
    ranges overlap on it."""
    if (
        not days_a or not days_b
        or start_a is None or start_b is None
        or end_a is None or end_b is None
    ):
        return False
    return bool(set(days_a) & set(days_b)) and start_a < end_b and start_b < end_a


@dataclass(frozen=True)
class TimeWindow:
    """A recurring weekly window, e.g. "Friday noon" or "no early TR"."""

    days: frozenset[str]
    start: datetime.time
    end: datetime.time
    reason: str = ""

    @classmethod
    def from_config(cls, raw: Mapping[str, object]) -> "TimeWindow":
        """``raw["between"]`` is a two-element ``[start, end]`` array --
        e.g. ``between = ["08:00", "09:00"]`` -- rather than separate
        ``start``/``end`` keys, since a bare "start"/"end" reads
        ambiguously next to a class's own actual start/end time (this is
        a *window* to match against, not one specific meeting)."""
        window_start, window_end = raw["between"]
        return cls(
            days=frozenset(raw.get("days", ())),
            start=record_utils.clock(window_start),
            end=record_utils.clock(window_end),
            reason=str(raw.get("reason", "")),
        )

    def overlaps(
        self,
        days: str | None,
        start: datetime.time | None,
        end: datetime.time | None,
    ) -> bool:
        return weekday_time_overlap(days, start, end, self.days, self.start, self.end)


@dataclass(frozen=True)
class PersonAlias:
    short: str
    subject: str | None = None


@dataclass(frozen=True)
class PersonRecord:
    """One persons.toml entry -- contractual facts about an instructor."""

    name: str
    max_load: float
    courses: tuple[str, ...] = ()
    aliases: tuple[PersonAlias, ...] = ()


@dataclass(frozen=True)
class ConstraintRule:
    """One hard rule using the same selectors as a preference rule."""

    direction: str = "+"
    name: str | None = None
    course: str | None = None
    section: str | None = None
    section_prefix: str | None = None
    room: str | tuple[str, ...] | None = None
    time: TimeWindow | None = None

    def __post_init__(self) -> None:
        if self.direction not in ("+", "-"):
            raise ValueError("constraint direction must be '+' or '-'")
        if self.name is None and self.room is None and self.time is None:
            raise ValueError(
                "a constraint rule requires name, room, and/or time"
            )

    def applies_to(self, course: str, section: str) -> bool:
        if self.course is not None and self.course != course:
            return False
        if self.section is not None and self.section != section:
            return False
        if (
            self.section_prefix is not None
            and not section.upper().startswith(self.section_prefix.upper())
        ):
            return False
        return True

    def allows(
        self,
        *,
        instructor: str,
        building: str,
        room: str,
        days: str | None,
        start: datetime.time | None,
        end: datetime.time | None,
        is_online: bool,
    ) -> bool:
        """Apply positive requirements or reject a forbidden combination."""
        matches: list[bool] = []
        if self.name is not None:
            matches.append(instructor == self.name)
        if self.room is not None:
            if is_online:
                if self.direction == "-":
                    return True
            else:
                matches.append(location_matches(building, room, self.rooms))
        if self.time is not None:
            if is_online:
                if self.direction == "-":
                    return True
            else:
                matches.append(self.time.overlaps(days, start, end))
        return all(matches) if self.direction == "+" else not all(matches)

    @property
    def rooms(self) -> tuple[str, ...]:
        if self.room is None:
            return ()
        return (self.room,) if isinstance(self.room, str) else self.room


@dataclass(frozen=True)
class PreferenceRule:
    """One normalized flat ``[[rules]]`` selector.

    The TOML loader uses a rule's explicit ``name`` to attach it to one
    ``PreferenceRecord``; rules without a name remain global. Comments and
    physical ordering in the file never determine scope. A TOML rule's
    positive/negative signed ``weight`` is normalized into ``direction`` and
    a non-negative magnitude here.

    ``course`` ("SUBJECT NUMBER", matching persons.toml's own
    convention), ``section``, ``section_prefix``, ``room``, and ``time``
    are all optional match keys -- an unset key matches anything, so a rule can be as
    broad ("this instructor generally avoids Corley") or as narrow ("this
    exact course-section must land in one of these rooms") as its fields
    specify. ``room`` may be one location or a tuple of alternatives;
    matching any alternative satisfies that selector once. ``section`` only makes sense alongside ``course`` -- a bare
    section code like "F01" repeats across unrelated courses. In contrast,
    ``section_prefix`` intentionally matches across courses, e.g. ``"TC"``.

    ``direction`` is ``"prefer"`` (subtracts ``weight`` from a matching
    candidate's cost -- a reward the solver seeks out) or ``"dislike"``
    (adds it -- a penalty the solver avoids); ``weight`` is always a
    non-negative magnitude, on the same 0-100 scale as the other soft
    costs. A weight of 100 is stronger than the 90-point under-load cost,
    but remains a soft objective rather than a hard constraint.
    """

    course: str | None = None
    section: str | None = None
    section_prefix: str | None = None
    room: str | tuple[str, ...] | None = None
    time: TimeWindow | None = None
    direction: str = "dislike"
    weight: float = 0.0

    def matches(
        self,
        *,
        course: str,
        section: str,
        building: str,
        room: str,
        days: str | None,
        start: datetime.time | None,
        end: datetime.time | None,
    ) -> bool:
        if self.course is not None and self.course != course:
            return False
        if self.section is not None and self.section != section:
            return False
        if (
            self.section_prefix is not None
            and not section.upper().startswith(self.section_prefix.upper())
        ):
            return False
        if self.room is not None and not location_matches(
            building, room, self.rooms
        ):
            return False
        if self.time is not None and not self.time.overlaps(days, start, end):
            return False
        return True

    @property
    def signed_weight(self) -> float:
        return -self.weight if self.direction == "prefer" else self.weight

    @property
    def rooms(self) -> tuple[str, ...]:
        if self.room is None:
            return ()
        return (self.room,) if isinstance(self.room, str) else self.room


@dataclass(frozen=True)
class PreferenceRecord:
    """One preferences.toml entry for the current term.

    ``allow_overload`` is this instructor's overload tolerance: ``True``
    means fine with it, ``False`` means avoid it (still soft -- see the
    package workload policy). The per-credit costs are supplied by
    constraints.toml.

    Flat named rules carry every selector-based preference and its weight.

    ``max_back_to_back`` caps how many consecutive same-day meetings this
    instructor tolerates before it's scored -- ``None`` means no cap
    (``allow_back_to_back`` alone decides). It only applies
    when ``allow_back_to_back`` is ``True``: back-to-back is already
    scored at every occurrence when it's ``False``, and a cap can't loosen
    that -- "no back-to-back at all" is a stricter statement than "no more
    than N in a row" ever needs to override. A run of exactly
    ``max_back_to_back`` meetings is fine; each meeting past that is its
    own scored finding using the configured back-to-back penalty, so a longer run
    costs more than a run of ``max_back_to_back + 1``.
    """

    name: str
    allow_overload: bool = True
    allow_back_to_back: bool = True
    max_back_to_back: int | None = None
    rules: tuple[PreferenceRule, ...] = ()

def load_persons(path: str | Path) -> dict[str, PersonRecord]:
    """Parse ``persons.toml`` into ``{name: PersonRecord}``."""
    with open(path, "rb") as handle:
        raw = PersonsFileSchema.model_validate(tomllib.load(handle))
    return {
        entry.name: PersonRecord(
            name=entry.name,
            max_load=entry.max_load,
            courses=tuple(entry.courses),
            aliases=tuple(
                PersonAlias(alias, None)
                if isinstance(alias, str)
                else PersonAlias(alias.short, alias.subject)
                for alias in entry.aliases
            ),
        )
        for entry in raw.persons
    }


def resolve_person_name(
    value: str, persons: Mapping[str, PersonRecord], *, subject: str | None = None
) -> str | None:
    """Resolve an exact person name or a configured, optionally scoped alias."""
    if value in persons:
        return value
    matches = {
        person.name
        for person in persons.values()
        for alias in person.aliases
        if alias.short == value and (alias.subject is None or alias.subject == subject)
    }
    if len(matches) > 1:
        raise ValueError(f"Ambiguous instructor alias {value!r}: {sorted(matches)}")
    return next(iter(matches), None)


def load_preferences(path: str | Path) -> dict[str, PreferenceRecord]:
    """Parse ``preferences.toml`` into ``{name: PreferenceRecord}``."""
    with open(path, "rb") as handle:
        raw = PreferencesFileSchema.model_validate(tomllib.load(handle))
    named_rules: dict[str, list[PreferenceRule]] = {}
    for rule in raw.rules:
        if rule.name is not None:
            named_rules.setdefault(rule.name, []).append(_parse_flat_rule(rule))
    return {
        entry.name: PreferenceRecord(
            name=entry.name,
            allow_overload=entry.allow_overload,
            allow_back_to_back=entry.allow_back_to_back,
            max_back_to_back=entry.max_back_to_back,
            rules=tuple(named_rules.get(entry.name, ())),
        )
        for entry in raw.instructors
    }


def load_global_rules(path: str | Path) -> tuple[PreferenceRule, ...]:
    """Parse flat rules without ``name``; these apply to every instructor."""
    with open(path, "rb") as handle:
        raw = PreferencesFileSchema.model_validate(tomllib.load(handle))
    return tuple(_parse_flat_rule(rule) for rule in raw.rules if rule.name is None)


def _parse_flat_rule(raw: FlatPreferenceRuleSchema) -> PreferenceRule:
    values = raw.model_dump(exclude={"name"}, exclude_none=True)
    values["direction"] = "prefer" if raw.weight > 0 else "dislike"
    values["weight"] = abs(raw.weight)
    return _parse_rule(values)


def _parse_rule(raw: Mapping[str, object]) -> PreferenceRule:
    direction = str(raw.get("direction", "dislike"))
    if direction not in ("prefer", "dislike"):
        raise ValueError(
            f"Rule direction must be 'prefer' or 'dislike', got {direction!r}"
        )
    section = raw.get("section")
    section_prefix = raw.get("section_prefix")
    if section is not None and "course" not in raw:
        raise ValueError("A rule's 'section' requires 'course' to also be set")
    raw_room = raw.get("room")
    return PreferenceRule(
        course=str(raw["course"]) if "course" in raw else None,
        section=str(section) if section is not None else None,
        section_prefix=(
            str(section_prefix).strip() if section_prefix is not None else None
        ),
        room=(
            str(raw_room) if isinstance(raw_room, str)
            else tuple(str(room) for room in raw_room)
            if raw_room is not None else None
        ),
        time=parse_rule_time(raw["time"]) if "time" in raw else None,
        direction=direction,
        weight=float(raw.get("weight", 0.0)),
    )


def parse_rule_time(raw: object) -> TimeWindow:
    if isinstance(raw, str):
        start, end = (part.strip() for part in raw.split("-", maxsplit=1))
        normalized = [part if ":" in part else f"{part}:00" for part in (start, end)]
        return TimeWindow.from_config(
            {"days": tuple("MTWRF"), "between": normalized}
        )
    if isinstance(raw, TimeWindowSchema):
        raw = raw.model_dump()
    if isinstance(raw, Mapping):
        return TimeWindow.from_config(raw)
    raise TypeError(f"Rule time must be a range string or table, got {type(raw).__name__}")


@dataclass(frozen=True)
class RecordReference:
    """Points at one row of one atomic class in a specific ``Schedule`` --
    what a web client needs to locate and highlight the record a finding
    is about, instead of guessing from free-text ``message`` (see
    docs/codes.md).

    ``class_index``/``record_index`` are positions in that ``Schedule``'s
    own ``classes``/``sections`` lists -- valid only for the serialized
    Schedule they were computed from, or its deterministic
    same-configuration flatten/regroup round trip (``to_records()`` then
    ``from_records()`` with the same relationships/catalogs -- grouping
    order is a function of each row's own identity fields, none of which
    a web edit can touch). They must never be persisted across schedule
    revisions. ``course_id`` is a cheap, self-describing companion a
    consumer can check the reference still points where it expects
    before trusting the indices.
    """

    class_index: int
    record_index: int
    course_id: str


def _indexed_sections(
    schedule: "Schedule",
) -> Iterator[tuple[RecordReference, Class, Section]]:
    """Every ``(reference, atomic class, section)`` triple in ``schedule``,
    in the same order ``_serialize_schedule`` (webapp.py) enumerates it --
    the one place that decides what a ``RecordReference``'s indices mean,
    so every ``check_*``/finding-producing function below shares it
    instead of each re-deriving its own ``enumerate`` (and risking a
    kind-specific trap like ``HybridClass.physical_section`` not actually
    being row 0).
    """
    for class_index, item in enumerate(schedule.classes):
        for record_index, section in enumerate(item.sections):
            yield (
                RecordReference(class_index, record_index, section.course_id),
                item, section,
            )


def _references_by_class(schedule: "Schedule") -> dict[int, tuple[RecordReference, ...]]:
    """Every class's own records as ``RecordReference``s, grouped by
    ``class_index`` -- derived from ``_indexed_sections`` so a "whole
    class" reference set (``check_atomic_class_rules``,
    ``_class_references_by_instructor``) never re-implements its own
    ``enumerate(item.sections)``.
    """
    grouped: dict[int, list[RecordReference]] = {}
    for ref, item, section in _indexed_sections(schedule):
        grouped.setdefault(ref.class_index, []).append(ref)
    return {index: tuple(refs) for index, refs in grouped.items()}


@dataclass(frozen=True)
class HardViolation:
    rule: str
    subject: str
    message: str
    references: tuple[RecordReference, ...] = ()


@dataclass(frozen=True)
class SoftFinding:
    rule: str
    instructor: str
    message: str
    penalty: float
    references: tuple[RecordReference, ...] = ()


def teaching_loads(schedule: "Schedule") -> dict[str, float]:
    """Authoritative total credit hours per instructor.

    A class's hours count once per distinct instructor teaching it, no matter
    how many CSV rows (sections) it has. Callers must not reproduce this from
    flattened records.
    """
    totals: dict[str, float] = {}
    for item in schedule:
        instructors = {s.instructor for s in item.sections if s.instructor}
        for instructor in instructors:
            totals[instructor] = totals.get(instructor, 0.0) + item.credit_hours
    return totals


@dataclass(frozen=True)
class InstructorLoadSummary:
    """One display/report row derived from authoritative atomic-class loads."""

    name: str
    hours: float
    target: float | None
    delta: float | None
    state: str
    position: str


def summarize_instructor_loads(
    loads: Mapping[str, float],
    persons: Mapping[str, PersonRecord],
    *,
    new_instructor_target: float,
    new_professor_target: float,
    overload_tolerance: float,
) -> tuple[InstructorLoadSummary, ...]:
    """Attach configured targets and status to shared teaching-load totals."""
    rows = []
    for name in sorted(set(persons) | set(loads), key=lambda value: value.lower()):
        if re.match(r"^(?:Staff|new_instructor)(?:\s|$)", name, re.I):
            target, position = new_instructor_target, "new_instructor"
        elif re.match(r"^new_professor(?:\s|$)", name, re.I):
            target, position = new_professor_target, "new_professor"
        else:
            target = persons[name].max_load if name in persons else None
            position = "instructor"
        hours = loads.get(name, 0.0)
        delta = hours - target if target is not None else None
        state = (
            "unknown" if target is None else
            "exact" if abs(delta) < 1e-9 else
            "under" if delta < 0 else
            "over" if delta <= overload_tolerance else "danger"
        )
        rows.append(InstructorLoadSummary(
            name=name, hours=hours, target=target, delta=delta,
            state=state, position=position,
        ))
    return tuple(rows)


def location_matches(building: str, room: str, locations: Iterable[str]) -> bool:
    """True if ``building``, ``room``, or "building room" combined is in
    ``locations`` -- accepts either form, matching how preferences.toml
    documents writing a location (e.g. "Corley" or "Corley 101")."""
    full = f"{building} {room}".strip()
    return bool({building, room, full} & set(locations))


def is_back_to_back(left: Section, right: Section) -> bool:
    if not (set(left.days or "") & set(right.days or "")):
        return False
    return left.end == right.start or right.end == left.start


_WEEKDAY_LETTERS = "MTWRF"


def _capped_back_to_back_findings(
    instructor: str, entries: list[tuple[RecordReference, Section]], cap: int,
    penalty: float,
) -> list[SoftFinding]:
    """Flag every meeting past the ``cap``-th in an instructor's own
    consecutive same-day run.

    Unlike ``is_back_to_back`` (any shared weekday between two possibly
    multi-day sections), a "run" here is built per single calendar day --
    the only way to walk a chain of 3+ meetings in order without a
    multi-day section like "MWF" ambiguously joining two unrelated runs.
    A run of exactly ``cap`` is unflagged; the (cap+1)-th meeting is one
    finding, the (cap+2)-th is another, and so on -- so a longer run
    always scores strictly more than a shorter one, same as stacking
    the configured penalty per over-cap join.

    A multi-day pattern like "MWF" walks the same run once per letter it
    spans (M, then W, then F) and finds the identical join each time --
    deduped by the course-id pair involved, so a genuine MWF run is one
    finding, not three, matching how the plain (uncapped) pairwise check
    already treats one back-to-back class-pair as a single finding
    regardless of how many shared weekdays it recurs on.

    ``entries`` carries each section's ``RecordReference`` alongside it
    (instead of a bare ``list[Section]``) because by the time a run is
    built here, the sections have been re-sorted/re-filtered per weekday
    -- there is no way to safely recover which atomic class/row a given
    ``Section`` came from after that without this.
    """
    findings: list[SoftFinding] = []
    # Keyed by (class_index, record_index) pairs, not course_id -- two
    # different physical meetings can share a course_id (FourCreditClass's
    # two rows are literally the same course/section), so a course_id key
    # could wrongly treat two distinct joins as the same one and drop the
    # second (see docs/codes.md).
    seen_joins: set[tuple[tuple[int, int], tuple[int, int]]] = set()
    for day in _WEEKDAY_LETTERS:
        day_entries = sorted(
            (
                (ref, s) for ref, s in entries
                if s.days and day in s.days and s.start is not None
            ),
            key=lambda pair: (pair[1].start.hour, pair[1].start.minute),
        )
        run: list[tuple[RecordReference, Section]] = []
        for entry in day_entries:
            ref, section = entry
            if run and run[-1][1].end == section.start:
                run.append(entry)
            else:
                run = [entry]
            if len(run) > cap:
                (prev_ref, prev), (last_ref, last) = run[-2], run[-1]
                join = (
                    (prev_ref.class_index, prev_ref.record_index),
                    (last_ref.class_index, last_ref.record_index),
                )
                if join in seen_joins:
                    continue
                seen_joins.add(join)
                findings.append(SoftFinding(
                    "back_to_back", instructor,
                    f"{instructor}: {prev.course_id} and "
                    f"{last.course_id} extend a same-day run past "
                    f"the {cap}-in-a-row cap",
                    penalty,
                    references=(prev_ref, last_ref),
                ))
    return findings


@dataclass(frozen=True)
class _OverloadStatus:
    instructor: str
    load: float
    max_load: float
    penalty: float


def _overload_statuses(
    schedule: "Schedule",
    persons: dict[str, PersonRecord],
    preferences: dict[str, PreferenceRecord],
    policy: WorkloadPolicySchema | None = None,
) -> list[_OverloadStatus]:
    policy = policy or WorkloadPolicySchema()
    """The single source of truth for "is this instructor overloaded".

    Anything within the configured overload tolerance of max_load isn't
    included at all -- it doesn't count as overload, so it's never
    reported by ``check_soft_preferences``. Everything this returns *is*
    overload, always soft -- ``penalty`` is ``preference.overload_penalty``
    per credit past the tolerance (``0.0`` when there is no preference
    preference record), plus ``OVERLOAD_FAR_PENALTY`` on top when they're also more than
    ``OVERLOAD_FAR_THRESHOLD`` credit hours over their own max_load *and*
    ``allow_overload`` -- see the module comment above
    the configured tolerance. Mirrors ``solver/constraints.py``'s load model
    exactly so the web UI's reported penalty matches what the solver
    actually optimized for.
    """
    statuses: list[_OverloadStatus] = []
    for instructor, load in sorted(teaching_loads(schedule).items()):
        person = persons.get(instructor)
        if person is None:
            continue
        excess = load - person.max_load
        if excess <= policy.overload_tolerance:
            continue
        preference = preferences.get(instructor)
        penalty = (
            (
                policy.penalties.permissive_overload_per_credit
                if preference.allow_overload
                else policy.penalties.strict_overload_per_credit
            ) * (excess - policy.overload_tolerance)
            if preference else 0.0
        )
        if (
            preference is not None and preference.allow_overload
            and excess > policy.far_overload_threshold
        ):
            penalty += policy.penalties.far_overload_extra
        statuses.append(_OverloadStatus(
            instructor=instructor,
            load=load,
            max_load=person.max_load,
            penalty=penalty,
        ))
    return statuses


def overlaps_in_time(left: Section, right: Section) -> bool:
    return weekday_time_overlap(
        left.days, left.start, left.end, right.days, right.start, right.end
    )


def check_conflicts(schedule: "Schedule") -> list[HardViolation]:
    """Report structural double-booking -- two different
    classes using the same room, or assigned to the same instructor, at an
    overlapping time.

    Independent of any config file -- this is a structural conflict in
    the schedule itself. A class's own multiple rows (FourCreditClass,
    HybridClass, CrossListingClass, CoreqClass) are never compared
    against each other here, since e.g. a genuine cross-listing is
    *meant* to share room/time/instructor -- it's one physical meeting
    filed under two catalog numbers, not a double-booking. Coreq/
    four-credit/hybrid scheduling legality doesn't need its own check
    here either; those values come from the workload policy.
    """
    entries = [
        (ref, item, section)
        for ref, item, section in _indexed_sections(schedule)
        if not section.is_online
    ]
    violations: list[HardViolation] = []
    for index, (ref_a, item_a, a) in enumerate(entries):
        for ref_b, item_b, b in entries[index + 1:]:
            if item_a is item_b or not overlaps_in_time(a, b):
                continue
            if a.room and a.building == b.building and a.room == b.room:
                location = f"{a.building} {a.room}".strip()
                violations.append(HardViolation(
                    "room_conflict", location,
                    f"{a.course_id} and {b.course_id} both use {location} "
                    f"at an overlapping time ({a.time_slot} / {b.time_slot})",
                    references=(ref_a, ref_b),
                ))
            if a.instructor and a.instructor == b.instructor:
                violations.append(HardViolation(
                    "instructor_conflict", a.instructor,
                    f"{a.course_id} and {b.course_id} both assign "
                    f"{a.instructor} at an overlapping time "
                    f"({a.time_slot} / {b.time_slot})",
                    references=(ref_a, ref_b),
                ))
    return violations


def check_atomic_class_rules(schedule: "Schedule") -> list[HardViolation]:
    """Report nonfatal construction issues that adjustment must repair.

    Sourced from each class's own ``schedule_issues``/``schedule_issue_rule``
    (all four two-row kinds -- see docs/codes.md) rather
    than a separate per-kind mapping kept here: any kind that carries
    ``schedule_issues`` is reported the same way, with no per-kind branch
    to remember to add.
    """
    violations: list[HardViolation] = []
    by_class = _references_by_class(schedule)
    for class_index, item in enumerate(schedule.classes):
        issues = getattr(item, "schedule_issues", ())
        if not issues:
            continue
        rule = item.schedule_issue_rule
        # The problem belongs to this atomic class as a whole -- every
        # record is listed, not just whichever row a message happens to
        # name. This is a "which class" pointer, not a claim that these
        # are the minimal faulty rows (see docs/codes.md).
        refs = by_class[class_index]
        violations.extend(
            HardViolation(rule, item.course_ids[0], message, references=refs)
            for message in issues
        )
    return violations


def check_constraint_rules(
    schedule: "Schedule", rules: Iterable[ConstraintRule],
) -> list[HardViolation]:
    """Validate mandatory instructor/room rules against grouped data."""
    configured = tuple(rules)
    violations: list[HardViolation] = []
    for ref, item, section in _indexed_sections(schedule):
        # A HybridClass's online companion row is derived, not a real
        # constraint target -- only its physical row is checked (matches
        # the pre-references behavior, but now via the shared iteration
        # helper instead of a second per-kind branch here).
        if isinstance(item, HybridClass) and section is not item.physical_section:
            continue
        course = f"{section.subject} {section.number}"
        for rule in configured:
            if not rule.applies_to(course, section.section):
                continue
            if rule.allows(
                instructor=section.instructor,
                building=section.building,
                room=section.room,
                days=section.days,
                start=section.start,
                end=section.end,
                is_online=section.is_online,
            ):
                continue
            conditions = []
            if rule.name is not None:
                conditions.append(f"name={rule.name!r}")
            if rule.room is not None:
                conditions.append(f"room={list(rule.rooms)!r}")
            if rule.time is not None:
                conditions.append("time=<configured window>")
            action = "must match" if rule.direction == "+" else "must not match"
            violations.append(HardViolation(
                "constraint_positive"
                if rule.direction == "+" else "constraint_negative",
                section.course_id,
                f"{section.course_id}: {action} "
                f"{', '.join(conditions)}",
                references=(ref,),
            ))
    return violations


def check_meeting_patterns(
    schedule: "Schedule", patterns: Iterable[MeetingPatternLike]
) -> list[HardViolation]:
    """Report physical sections outside their configured pattern domain."""
    configured = tuple(patterns)
    if not configured:
        return []
    violations: list[HardViolation] = []
    for ref, item, section in _indexed_sections(schedule):
        if section.is_online or matches_configured_pattern(
            item, section, configured
        ):
            continue
        violations.append(HardViolation(
            "meeting_pattern",
            section.course_id,
            f"{section.course_id} uses an unconfigured meeting pattern: "
            f"{section.time_slot} ({section.duration} minutes)",
            references=(ref,),
        ))
    return violations


def _class_references_by_instructor(
    schedule: "Schedule",
) -> dict[str, tuple[RecordReference, ...]]:
    """Every record of every atomic class an instructor appears in.

    Deliberately not "just the rows that literally name them" --
    ``teaching_loads()`` credits an instructor a class's full
    ``credit_hours`` the moment *any* row of that class names them (see
    docs/codes.md), so overload/under_load references have to cover the
    same set of records that definition draws from, or a CrossListingClass
    with two different instructors would silently lose the row that
    doesn't happen to name the one a finding is about.
    """
    by_class = _references_by_class(schedule)
    by_instructor: dict[str, list[RecordReference]] = {}
    for class_index, item in enumerate(schedule.classes):
        for instructor in {s.instructor for s in item.sections if s.instructor}:
            by_instructor.setdefault(instructor, []).extend(by_class[class_index])
    return {name: tuple(refs) for name, refs in by_instructor.items()}


def check_soft_preferences(
    schedule: "Schedule",
    preferences: dict[str, PreferenceRecord],
    persons: dict[str, PersonRecord],
    global_rules: tuple[PreferenceRule, ...] = (),
    workload_policy: WorkloadPolicySchema | None = None,
    back_to_back_policy: BackToBackPolicySchema | None = None,
) -> tuple[float, list[SoftFinding]]:
    """Score a schedule against preferences.toml.

    Returns ``(total_penalty, findings)`` -- 0 means every preference was
    honored, lower is always better. Every rule here is soft, including
    ``max_load`` according to the configured workload policy. Being outside
    a ``prefer`` rule is not
    reported as a violation: matching candidates receive their configured
    weighted reward in the solver, while absence of a match adds no finding.

    ``global_rules`` plus each matching instructor's own ``rules`` (see
    ``PreferenceRule``) are checked too, but only their ``"dislike"``
    side -- a matching ``"prefer"`` rule still steers the solver (it's
    scored in ``solver/candidates.py``'s ``preference_cost``) but isn't reported
    here, since a *satisfied* preference isn't a violation to surface
    next to everything else this function returns.
    """
    workload_policy = workload_policy or WorkloadPolicySchema()
    back_to_back_policy = back_to_back_policy or BackToBackPolicySchema()
    class_refs = _class_references_by_instructor(schedule)
    findings: list[SoftFinding] = [
        SoftFinding(
            "overload", status.instructor,
            f"{status.instructor}: {status.load:g} credit hours exceeds "
            f"max_load {status.max_load:g}",
            status.penalty,
            references=class_refs.get(status.instructor, ()),
        )
        for status in _overload_statuses(
            schedule, persons, preferences, workload_policy,
        )
    ]

    loads = teaching_loads(schedule)
    for instructor, person in sorted(persons.items()):
        load = loads.get(instructor, 0.0)
        deficit = person.max_load - load
        if deficit > 0:
            findings.append(SoftFinding(
                "under_load", instructor,
                f"{instructor}: {load:g} credit hours is under max_load "
                f"{person.max_load:g}",
                deficit * workload_policy.penalties.underload_per_credit,
                # Legitimately empty for an instructor currently teaching
                # nothing (see docs/codes.md) -- the web UI still falls
                # back to a plain instructor-tab link via `subject` then.
                references=class_refs.get(instructor, ()),
            ))

    sections = [
        (ref, section)
        for ref, item, section in _indexed_sections(schedule)
        if section.instructor
    ]

    for ref, section in sections:
        preference = preferences.get(section.instructor)
        course = f"{section.subject} {section.number}"
        applicable_rules = list(global_rules)
        if preference is not None:
            applicable_rules.extend(preference.rules)
        for rule in applicable_rules:
            if rule.direction != "dislike":
                continue
            if rule.matches(
                course=course, section=section.section,
                building=section.building, room=section.room,
                days=section.days, start=section.start, end=section.end,
            ):
                findings.append(SoftFinding(
                    "custom_rule", section.instructor,
                    f"{section.course_id}: matches a custom dislike rule "
                    f"(weight {rule.weight:g})",
                    rule.weight,
                    references=(ref,),
                ))
        if preference is None:
            continue
    by_instructor: dict[str, list[tuple[RecordReference, Section]]] = {}
    for ref, section in sections:
        if not section.is_online:
            by_instructor.setdefault(section.instructor, []).append((ref, section))
    for instructor, instructor_entries in by_instructor.items():
        preference = preferences.get(instructor)
        if preference is None:
            continue
        if not preference.allow_back_to_back:
            for i, (ref_left, left) in enumerate(instructor_entries):
                for ref_right, right in instructor_entries[i + 1:]:
                    if is_back_to_back(left, right):
                        findings.append(SoftFinding(
                            "back_to_back", instructor,
                            f"{instructor}: {left.course_id} and "
                            f"{right.course_id} are back-to-back",
                            back_to_back_policy.penalty,
                            references=(ref_left, ref_right),
                        ))
        elif preference.max_back_to_back is not None:
            findings.extend(_capped_back_to_back_findings(
                instructor, instructor_entries, preference.max_back_to_back,
                back_to_back_policy.penalty,
            ))

    total = sum(finding.penalty for finding in findings)
    return total, findings


def check_workload_hard_caps(
    schedule: "Schedule",
    persons: dict[str, PersonRecord],
    workload_policy: WorkloadPolicySchema | None = None,
    new_instructor_policy: NewInstructorPolicySchema | None = None,
    new_professor_policy: NewProfessorPolicySchema | None = None,
) -> list[HardViolation]:
    """Report loads the solver would never actually produce: a hard cap
    a configured instructor's ``hard_load_cap_tolerance`` allows no
    further leeway past, or a New Instructor/New Professor identity's
    contract load (which the solver enforces with *no* tolerance at all,
    unlike a configured person -- see ``add_load_terms``). Distinct rules
    (``hard_load_cap`` vs ``new_hire_contract_load``) since they're
    different caps for different reasons, even though both come from the
    same underlying per-instructor totals.

    Uses ``teaching_loads()``/``_class_references_by_instructor()`` --
    the same "every distinct instructor named anywhere in a class counts
    it once" definition ``overload``/``under_load`` already use. This
    used to be deliberately narrower (see the removed
    ``_primary_section_instructor_loads``, "Plan A" in docs/codes.md),
    matching what ``solver/constraints.py``'s ``add_load_terms`` used to
    attribute -- only the class's first row's instructor. Now that
    ``add_load_terms`` itself counts every row (see docs/codes.md), the
    two definitions are simply the same one; keeping a second, narrower
    copy here would reintroduce the exact report/solver disagreement this
    whole design exists to prevent.
    """
    workload_policy = workload_policy or WorkloadPolicySchema()
    new_instructor_policy = new_instructor_policy or NewInstructorPolicySchema()
    new_professor_policy = new_professor_policy or NewProfessorPolicySchema()
    totals = teaching_loads(schedule)
    references = _class_references_by_instructor(schedule)
    violations: list[HardViolation] = []
    for instructor, load in sorted(totals.items()):
        if is_new_instructor(instructor):
            cap, rule = new_instructor_policy.contract_load, "new_hire_contract_load"
        elif is_new_professor(instructor):
            cap, rule = new_professor_policy.contract_load, "new_hire_contract_load"
        else:
            person = persons.get(instructor)
            if person is None:
                continue
            cap = person.max_load + workload_policy.hard_load_cap_tolerance
            rule = "hard_load_cap"
        if load > cap:
            violations.append(HardViolation(
                rule, instructor,
                f"{instructor}: {load:g} credit hours exceeds the hard "
                f"cap of {cap:g}",
                references=references.get(instructor, ()),
            ))
    return violations


def check_new_hire_counts(
    schedule: "Schedule",
    new_instructor_policy: NewInstructorPolicySchema | None = None,
    new_professor_policy: NewProfessorPolicySchema | None = None,
) -> list[HardViolation]:
    """Report when the number of *distinct* New Instructor/New Professor
    identities actually in use falls outside ``allowed_counts``.

    Counts identities directly from the current rows -- not the
    solver's own ``used`` CP-SAT variables (``add_placeholder_count_terms``),
    which additionally assume contiguous use (identity 2 only used if 1
    already is); a raw/manually-edited schedule doesn't have to satisfy
    that invariant, so re-deriving it here would risk under- or
    over-counting. ``references`` legitimately comes back empty when the
    count itself is the problem at zero (``allowed_counts = [1]`` but
    none currently in use).
    """
    new_instructor_policy = new_instructor_policy or NewInstructorPolicySchema()
    new_professor_policy = new_professor_policy or NewProfessorPolicySchema()
    violations: list[HardViolation] = []
    for label, rule, is_kind, policy in (
        ("new_instructor", "new_instructor_count", is_new_instructor, new_instructor_policy),
        ("new_professor", "new_professor_count", is_new_professor, new_professor_policy),
    ):
        used: set[str] = set()
        refs: list[RecordReference] = []
        for ref, item, section in _indexed_sections(schedule):
            if section.instructor and is_kind(section.instructor):
                used.add(section.instructor)
                refs.append(ref)
        count = len(used)
        if count not in policy.allowed_counts:
            violations.append(HardViolation(
                rule, label,
                f"{count} distinct {label} identities are in use; "
                f"allowed counts are {policy.allowed_counts}",
                references=tuple(refs),
            ))
    return violations


@dataclass(frozen=True)
class ScheduleEvaluation:
    """All deterministic statistics derived from one grouped schedule."""

    atomic_classes: int
    row_count: int
    loads: dict[str, float]
    hard_violations: tuple[HardViolation, ...]
    soft_penalty: float
    soft_findings: tuple[SoftFinding, ...]


def evaluate_schedule(
    schedule: "Schedule",
    preferences: dict[str, PreferenceRecord],
    persons: dict[str, PersonRecord],
    global_rules: tuple[PreferenceRule, ...] = (),
    meeting_patterns: Iterable[MeetingPatternLike] = (),
    constraint_rules: Iterable[ConstraintRule] = (),
    workload_policy: WorkloadPolicySchema | None = None,
    back_to_back_policy: BackToBackPolicySchema | None = None,
    new_instructor_policy: NewInstructorPolicySchema | None = None,
    new_professor_policy: NewProfessorPolicySchema | None = None,
) -> ScheduleEvaluation:
    """Evaluate only domain objects; raw CSV rows are not accepted here."""
    soft_penalty, soft_findings = check_soft_preferences(
        schedule, preferences, persons, global_rules,
        workload_policy, back_to_back_policy,
    )
    return ScheduleEvaluation(
        atomic_classes=len(schedule),
        row_count=len(schedule.to_records()),
        loads=teaching_loads(schedule),
        hard_violations=tuple(
            check_atomic_class_rules(schedule)
            + check_conflicts(schedule)
            + check_meeting_patterns(schedule, meeting_patterns)
            + check_constraint_rules(schedule, constraint_rules)
            + check_workload_hard_caps(
                schedule, persons, workload_policy,
                new_instructor_policy, new_professor_policy,
            )
            + check_new_hire_counts(
                schedule, new_instructor_policy, new_professor_policy,
            )
        ),
        soft_penalty=soft_penalty,
        soft_findings=tuple(soft_findings),
    )


# ---- Excel workbook builders ----
#
# The weekly views work directly from grouped Class/Section objects, so
# companion rows can be distinguished from genuine scheduling conflicts.

_WEEKDAYS = (
    ("M", "Monday"), ("T", "Tuesday"), ("W", "Wednesday"),
    ("R", "Thursday"), ("F", "Friday"),
)


def _write_raw_excel(df: pd.DataFrame, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Schedule", index=False)
        if df.empty:
            return
        ws = writer.book["Schedule"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center")
        for column in ws.columns:
            width = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column
            )
            ws.column_dimensions[column[0].column_letter].width = min(
                35, max(10, width + 2)
            )


def _weekly_workbook(schedule: "Schedule", *, group: str) -> Workbook:
    """``group`` is ``"instructor"`` or ``"room"``."""
    # Keep each section paired with the atomic class (``Class``) it came
    # from -- ``_build_weekly_sheet`` needs that to tell a real
    # double-booking from a HybridClass/CrossListingClass companion
    # meeting that's *meant* to land on the same slot.
    entries = [(item, section) for item in schedule for section in item.sections]

    def key_of(section: Section) -> str:
        if group == "instructor":
            return section.instructor
        return f"{section.building} {section.room}".strip()

    groups: dict[str, list[tuple[Class, Section]]] = {}
    for item, section in entries:
        if group == "room" and section.is_online:
            continue  # no physical meeting to place on a room grid
        key = key_of(section)
        if not key:
            continue
        groups.setdefault(key, []).append((item, section))

    workbook = Workbook()
    workbook.remove(workbook.active)
    if not groups:
        ws = workbook.create_sheet("No scheduled data")
        ws["A1"] = f"No rows with {group} assignments."
        return workbook

    used_titles: set[str] = set()
    for resource in sorted(groups):
        title = _safe_sheet_title(resource, used_titles)
        ws = workbook.create_sheet(title)
        _build_weekly_sheet(ws, resource, groups[resource], group)
    return workbook


def _merged_anchor(ws, row: int, column: int):
    """The writable top-left cell for ``(row, column)`` -- itself if it
    isn't part of a merge, otherwise that merged range's own anchor cell
    (the only one of its cells whose ``.value`` isn't read-only)."""
    cell = ws.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return cell  # pragma: no cover -- a MergedCell always belongs to some range


def _build_weekly_sheet(
    ws, resource: str, entries: list[tuple[Class, Section]], group: str
) -> None:
    # Restart the ten-colour palette for every instructor/room sheet.
    # Colour belongs to the atomic Class object, so all of its component
    # sections and meeting rows stay visually tied together.
    color_by_class: dict[int, str] = {}
    for item, _section in entries:
        class_key = id(item)
        if class_key not in color_by_class:
            color_by_class[class_key] = _ATOMIC_CLASS_COLORS[
                len(color_by_class) % len(_ATOMIC_CLASS_COLORS)
            ]

    physical = [(item, s) for item, s in entries if not s.is_online]
    starts = [_minutes(s.start) for _, s in physical if s.start is not None]
    ends = [_minutes(s.end) for _, s in physical if s.end is not None]
    first_minute = min([8 * 60, *(starts or [8 * 60])])
    last_minute = max([17 * 60, *(ends or [17 * 60])])
    # Meeting blocks round up to the half-hour grid: 50 minutes occupies
    # 60 visual minutes, 80 occupies 90.
    first_minute = first_minute // 30 * 30
    last_minute = (last_minute + 29) // 30 * 30

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.merge_cells("A1:F1")
    resource_label = "Instructor" if group == "instructor" else "Room"
    ws["A1"] = f"{resource_label} Schedule -- {resource}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Time", *(day_label for _, day_label in _WEEKDAYS)]
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(2, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 13
    for letter in "BCDEF":
        ws.column_dimensions[letter].width = 23

    slots = list(range(first_minute, last_minute, 30))
    light_border = Border(
        bottom=Side(style="hair", color="D9E2F3"),
        right=Side(style="hair", color="D9E2F3"),
    )
    for offset, minute in enumerate(slots, start=3):
        ws.cell(offset, 1, _clock(minute))
        ws.cell(offset, 1).alignment = Alignment(horizontal="right", vertical="top")
        ws.row_dimensions[offset].height = 18
        for column in range(1, 7):
            ws.cell(offset, column).border = light_border

    # Cells are claimed per atomic class, not per raw section -- a
    # HybridClass/CrossListingClass companion meeting landing on the same
    # slot as its own sibling is expected by design (one physical meeting
    # filed under two rows), not a double-booking. Only a slot already
    # claimed by a *different* class is a real conflict.
    occupied: dict[tuple[int, int], int] = {}
    ordered = sorted(
        physical, key=lambda pair: (pair[1].start or datetime.time(0, 0), pair[1].course_id)
    )
    for item, section in ordered:
        secondary = (
            f"{section.building} {section.room}".strip()
            if group == "instructor" else section.instructor
        )
        text = f"{section.course_id}\n{secondary or ''}\n{_clock_range(section)}"
        fill = color_by_class[id(item)]
        start_row = 3 + (_minutes(section.start) - first_minute) // 30
        duration = _minutes(section.end) - _minutes(section.start)
        visual_slots = max(1, (duration + 29) // 30)
        end_row = start_row + visual_slots - 1
        for day, _day_label in _WEEKDAYS:
            if day not in (section.days or ""):
                continue
            column = 2 + [w[0] for w in _WEEKDAYS].index(day)
            cells = {(r, column) for r in range(start_row, end_row + 1)}
            occupants = {occupied[c] for c in cells if c in occupied}
            if occupants == {id(item)}:
                # The same atomic class's own companion meeting -- merge
                # into the existing block instead of flagging anything.
                # start_row is this meeting's *own* row, which can land
                # inside (not at the top of) an already-merged block --
                # e.g. an 80-minute meeting rounds up to a 90-visual-minute
                # block (see visual_slots above), so a back-to-back
                # companion meeting immediately after it starts at a row
                # already swallowed by that rounding. ws.cell() on such a
                # row returns a read-only MergedCell, so resolve to the
                # block's real top-left anchor first.
                anchor = _merged_anchor(ws, start_row, column)
                if section.course_id not in (anchor.value or ""):
                    anchor.value = f"{anchor.value}\n/ {section.course_id}".strip()
                occupied |= {cell: id(item) for cell in cells}
                continue
            if occupants:
                # A genuinely different class landed on the same cell --
                # a real scheduling conflict. Keep both visible rather
                # than silently overwriting one: append to the existing
                # cell and flag it red instead of creating a second block.
                # Same MergedCell hazard as above.
                anchor = _merged_anchor(ws, start_row, column)
                anchor.value = f"{anchor.value or ''}\nCONFLICT: {text}".strip()
                anchor.fill = PatternFill("solid", fgColor="FF0000")
                anchor.font = Font(size=9, bold=True, color="FFFFFF")
                continue
            occupied |= {cell: id(item) for cell in cells}
            ws.merge_cells(
                start_row=start_row, start_column=column,
                end_row=end_row, end_column=column,
            )
            cell = ws.cell(start_row, column, text)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(size=9, color="1F1F1F")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin", color="7F8C8D"),
                right=Side(style="thin", color="7F8C8D"),
                top=Side(style="thin", color="7F8C8D"),
                bottom=Side(style="thin", color="7F8C8D"),
            )

    if group == "instructor":
        unscheduled = [s for _, s in entries if s.is_online]
        if unscheduled:
            row_number = 3 + len(slots) + 2
            ws.cell(row_number, 1, "Online / TBA")
            ws.cell(row_number, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row_number, 1).fill = PatternFill("solid", fgColor="7F8C8D")
            ws.merge_cells(
                start_row=row_number, start_column=2,
                end_row=row_number, end_column=6,
            )
            ws.cell(row_number, 2, ", ".join(s.course_id for s in unscheduled))
            ws.cell(row_number, 2).alignment = Alignment(wrap_text=True)

    ws.auto_filter.ref = f"A2:F{2 + len(slots)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"


def _safe_sheet_title(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", value).strip("'") or "Schedule"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate.casefold() in used:
        suffix = f" ({counter})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate.casefold())
    return candidate


_ATOMIC_CLASS_COLORS = (
    "E6F3FF",  # Light blue
    "E6FFE6",  # Light green
    "FFF2E6",  # Light orange
    "F0E6FF",  # Light purple
    "FFE6F2",  # Light pink
    "E6FFFF",  # Light cyan
    "FFFACD",  # Light yellow
    "F5F5DC",  # Light beige
    "E6E6FA",  # Light lavender
    "F0F8E6",  # Light mint
)


def _minutes(value: datetime.time) -> int:
    return value.hour * 60 + value.minute


def _clock(minutes: int) -> str:
    hour, minute = divmod(minutes, 60)
    suffix = "AM" if hour < 12 else "PM"
    shown = hour % 12 or 12
    return f"{shown}:{minute:02d} {suffix}"


def _clock_range(section: Section) -> str:
    if section.start is None or section.end is None:
        return ""
    return f"{_clock(_minutes(section.start))}-{_clock(_minutes(section.end))}"
